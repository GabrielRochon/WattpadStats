# Basic web scraper that collects views, votes & parts of your Wattpad story
# Libraries
from bs4 import BeautifulSoup
from urllib.request import Request, urlopen
from openpyxl import Workbook, load_workbook
from datetime import datetime
import os
import random
import time

try:
    from openpyxl.cell import get_column_letter
except ImportError:
    from openpyxl.utils import get_column_letter

# Source files
from slicing import *

# ENTER YOUR STORY INFO HERE ##################################################
url = 'https://www.wattpad.com/story/226573279-virtuel'
random_delay = True    # RECOMMENDED TO STAY TRUE FOR IP BLACKLISTING
###############################################################################

print('###########################################')
print('#     WattpadStats, by Gabriel Rochon     #')
print('###########################################\n')

# MAIN PAGE ###################################################################
# Extract stats from your story page's HTML
req = Request(url, headers={'User-Agent': 'Mozilla/5.0'})     # So bots don't 403 you
webpage = urlopen(req).read()
soup = BeautifulSoup(webpage, 'html.parser')
stats = soup.findChild("div", {"class": "meta"})

leaves = list()
for i in stats.children:
    if(i != '\n'):
        if(len(leaves) < 2):
            leaves.append(i['title'])
        else:
            leaves.append(i.contents[0])

title_wrapper = soup.findChild("div", {"class": "container"})
title = title_wrapper.findChild("h1").contents[0][:-1]

# EVERY CHAPTER ###############################################################
# Get every chapter's URL and grab individual stats for each one of them
urls = list()
chapters = soup.find("ul", {"class": "table-of-contents"})
chapters_list = chapters.findAll("a")
for i in chapters_list:
    urls.append(i['href'])

# Access every url to retrieve individual chapter stats
# WARNING: Gotta be careful with this, as Wattpad may see this as a potential DDoS
#   This process will take some time to ensure Wattpad doesn't blacklist our IP address
#   while accessing all of the stories' urls
print('Reading chapters, delaying reads so Wattpad doesn\'t block http requests.')
print('This will take around ' + str(len(chapters_list)*3) + ' seconds.')
print('(NOT RECOMMENDED) If you wish to speed up the process, remove the random delay option in scraper.py.\n')

init_time = time.time()

chap_stats_list = list()
count = 1
for chap_url in urls:
    print('Reading chapter ' + str(count) + '...', end = ' ', flush = True)

    # Get info
    chap_req = Request('http://wattpad.com/' + chap_url, headers={'User-Agent': 'Mozilla/5.0'})     # So bots don't 403 you
    chap_webpage = urlopen(chap_req).read()
    chap_soup = BeautifulSoup(chap_webpage, 'html.parser')

    chap_reads = chap_soup.findChild("span", {"class": "reads"})
    chap_votes = chap_soup.findChild("span", {"class": "votes"})
    chap_comments = chap_soup.findChild("span", {"class": "comments on-comments"}).findChild("a")
    chap_title = chap_soup.findChild("header", {"class": "panel panel-reading text-center"}).findChild("h2").contents

    chap_stats = [int(chap_reads.contents[2]), int(chap_votes.contents[2]), int(chap_comments.contents[0]), chap_title[0]]

    chap_stats_list.append(chap_stats)

    if random_delay:
        delay = random.random()*2 + 1            # Random delay from 1 to 3 secs
        time.sleep(delay)
    print('OK!')
    count += 1

print('\n')
###############################################################################

# OUTPUT TO EXCEL FILE ########################################################
# Slice those strings into numbers
reads = slice_reads(leaves[0])
votes = slice_votes(leaves[1])
parts = slice_parts(leaves[2])

# Populate Excel file with numerical stats
filename = title + ' stats.xlsx'
existing_file = False
if(os.path.isfile(filename)):
    existing_file = True

row = 1
column = 1

inc_reads = 0
inc_votes = 0
inc_parts = 0

# FIRST WORKSHEET: GENERAL INFO
if not existing_file:
    wb = Workbook()
    ws = wb.active
    ws.title = 'General'
    header = [ "DATE", "READS", "", "VOTES", "", "PARTS", "" ]
    ws.append(header)
    row += 1
else:
    wb = load_workbook(filename=filename)
    ws = wb['General']
    row = ws.max_row + 1

    # Increments from yesterday
    yesterday = ws[ws.max_row]
    inc_reads = int(reads) - int(yesterday[1].value)
    inc_votes = int(votes) - int(yesterday[3].value)
    inc_parts = int(parts) - int(yesterday[5].value)

# Sets column sizes
columns = ['A', 'B', 'C', 'D', 'E', 'F', 'G']       # Pretty sure there's a better way to do it, but for now it's fine
widths = [20, 10, 5, 10, 5, 10, 5]
for (width, col) in zip(widths, columns):
    ws.column_dimensions[col].width = width

# Actual output
today = datetime.today().strftime('%d/%m/%Y')
data = [ today, reads, inc_reads, votes, inc_votes, parts, inc_parts ]

for item in data:
    if column == 3 or column == 5 or column == 7:
        if item == 0:
            ws.cell(row, column, '')        # If incr. is null, don't output it
        else:
            ws.cell(row, column, '(+' + str(item) + ')')
    else:
        if column == 2 or column == 4 or column == 6:
            ws.cell(row, column, int(item))
        else:
            ws.cell(row, column, item)
    column += 1

# READS/VOTES/PARTS WORKSHEETS ################################################
sheet_titles = ['Reads', 'Votes', 'Comments']

for i in range(3):

    if not existing_file:
        wb.create_sheet(sheet_titles[i] + ' per chap')
    ws = wb[sheet_titles[i] + ' per chap']

    row = 1
    column = 1

    chap_inc = list()

    ws.cell(row, column, 'DATE')
    ws.column_dimensions['A'].width = 20
    ws.merge_cells(start_row=row, start_column=column, end_row=row+1, end_column=column)
    column += 1

    for chap in chap_stats_list:
        ws.cell(row, column, 'CHAP #' + str(int((column+1)/2)))
        ws.merge_cells(start_row=row, start_column=column, end_row=row, end_column=column+1)
        ws.cell(row+1, column, chap[3])
        ws.merge_cells(start_row=row+1, start_column=column, end_row=row+1, end_column=column+1)
        column += 2
    column = 1

    row = ws.max_row + 1
    ws.cell(row, column, today)
    column += 1

    for chap in chap_stats_list:
        ws.cell(row, column, chap[i])
        column += 2

    # Increments
    column = 3
    if existing_file:
        for chap in chap_stats_list:
            yesterday = ws.cell(row=row-1, column=column-1).value
            if yesterday is not None:
                chap_inc = chap[i] - yesterday
                if chap_inc > 0:
                    ws.cell(row, column, '(+' + str(chap_inc) + ')')
                column += 2

    # Resize column dimensions
    for i in range(ws.max_column):
        i = i+1
        if i == 1:
            width = 20
        else:
            if i % 2 == 0:
                width = 10
            else:
                width = 5

        col = get_column_letter(i)
        ws.column_dimensions[col].width = width

###############################################################################

print('Scraping ended in ' + str(round(time.time()-init_time,2)) + ' secs.')
print('Thank you for using WattpadStats! :)\n')
print('(c) Gabriel Rochon')
wb.save(filename)
